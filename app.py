import os, io
from flask import Flask, render_template, request, jsonify, send_file, flash, redirect, url_for, session
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
from models import Database

# ✅ tout est à la racine (HTML + CSS)
app = Flask(__name__, template_folder=".", static_folder=".")
app.secret_key = os.getenv('SECRET_KEY', 'ecole_mont_sion_secret_key')
db = Database()

MOTS_DE_PASSE = ['kouame', 'arrow', 'celestin', 'Viviane']
MATIERES = ['Mathématiques','Communication écrite','Lecture','Anglais','SVT','Histoire-géographie','Espagnol','EPS','Conduite','Physique']

# ---------- AUTH ----------
@app.route('/auth')
def auth():
    nxt = request.args.get('next', 'accueil')
    return render_template('auth.html', next_page=nxt)

@app.route('/verifier', methods=['POST'])
def verifier():
    pwd = request.form.get('password', '').strip()
    nxt = request.form.get('next_page', 'accueil')
    if pwd in MOTS_DE_PASSE:
        session['access_granted'] = True
        return redirect(url_for(nxt))
    flash('❌ Accès refusé – contactez la direction.', 'error')
    return redirect(url_for('auth', next=nxt))

def require_auth(f):
    def wrapped(*a, **k):
        if not session.get('access_granted'):
            return redirect(url_for('auth', next=request.endpoint))
        return f(*a, **k)
    wrapped.__name__ = f.__name__
    return wrapped

# ---------- ACCUEIL ----------
@app.route('/')
def accueil():
    return render_template('accueil.html')

# ---------- INSCRIPTION ----------
@app.route('/inscription')
def inscription():
    return render_template('inscription.html')

@app.route('/inscrire_ecolier', methods=['POST'])
def inscrire_ecolier():
    data = request.json
    if not is_valid_phone(data.get('telephone')):
        return jsonify(success=False, error="Téléphone doit commencer par 01 et contenir 10 chiffres"), 400
    db.add_ecolier(data)
    return jsonify(success=True)

@app.route('/inscrire_eleve', methods=['POST'])
def inscrire_eleve():
    data = request.json
    if not is_valid_phone(data.get('telephone')):
        return jsonify(success=False, error="Téléphone doit commencer par 01 et contenir 10 chiffres"), 400
    db.add_eleve(data)
    return jsonify(success=True)

def is_valid_phone(tel: str) -> bool:
    return bool(tel) and tel.isdigit() and len(tel) == 10 and tel.startswith('01')

# ---------- LISTES ----------
@app.route('/liste_eleves')
def liste_eleves():
    return render_template('liste_eleves.html', eleves=db.get_eleves())

@app.route('/liste_ecoliers')
def liste_ecoliers():
    return render_template('liste_ecoliers.html', ecoliers=db.get_ecoliers())

# ---------- SCOLARITÉ ----------
@app.route('/scolarite')
@require_auth
def scolarite():
    students = db.get_all()
    for s in students:
        try:
            montant = int(str(s.get('montant_scolarite', '0')).strip())
        except ValueError:
            montant = 0
        total = db.get_total_paid(s)
        s['total_paid'] = total
        s['reste'] = montant - total
    return render_template('scolarite.html', students=students, ecoliers=db.get_ecoliers())

@app.route('/paiement', methods=['POST'])
def paiement():
    data = request.json
    ok = db.add_payment(data['student_id'], data['student_type'], data['amount'])
    if ok:
        st = None
        if data['student_type'] == 'ecolier':
            st = next((x for x in db.get_ecoliers() if x['id'] == data['student_id']), None)
        else:
            st = next((x for x in db.get_eleves() if x['id'] == data['student_id']), None)
        if st:
            try:
                montant = int(str(st.get('montant_scolarite', '0')).strip())
            except ValueError:
                montant = 0
            total = db.get_total_paid(st)
            return jsonify(success=True, total_paid=total, reste=montant - total)
    return jsonify(success=False)

# ---------- NOTES ----------
@app.route('/notes')
@require_auth
def notes():
    ecoliers = db.get_ecoliers(); eleves = db.get_eleves()
    classes_ecoliers = sorted(set(e['classe'] for e in ecoliers))
    classes_eleves = sorted(set(e['classe'] for e in eleves))
    return render_template('notes.html', classes_ecoliers=classes_ecoliers, classes_eleves=classes_eleves, matieres=MATIERES)

@app.route('/get_students_by_class', methods=['POST'])
def get_students_by_class():
    d = request.json
    students = [s for s in (db.get_ecoliers() if d['is_ecolier'] else db.get_eleves()) if s['classe'] == d['classe']]
    return jsonify(students=students)

@app.route('/save_notes', methods=['POST'])
def save_notes():
    for n in request.json['notes']:
        note = float(n['note'])
        if not (0 <= note <= 99):
            return jsonify(success=False, error="Note doit être entre 0 et 99"), 400
        db.add_note(n['student_id'], n['student_type'], n['classe'], n['matiere'], note)
    return jsonify(success=True)

@app.route('/vue_notes')
@require_auth
def vue_notes():
    notes = db.get_notes()
    classes = sorted(set(n['classe'] for n in notes))
    matieres = sorted(set(n['matiere'] for n in notes))
    return render_template('vue_notes.html', classes=classes, matieres=matieres)

@app.route('/get_all_notes', methods=['POST'])
def get_all_notes():
    classe = request.json.get('classe', '')
    matiere = request.json.get('matiere', '')
    notes = db.get_notes(classe=classe or None, matiere=matiere or None)
    all_st = {s['id']: f"{s['nom']} {s['prenoms']}" for s in db.get_all()}
    out = [{'student_name': all_st.get(n['student_id'], 'Inconnu'), 'classe': n['classe'], 'matiere': n['matiere'], 'note': n['note'], 'date': n['date']} for n in notes]
    return jsonify(notes=out)

# ---------- MOYENNES ----------
@app.route('/moyenne')
@require_auth
def moyenne():
    ecoliers = db.get_ecoliers(); eleves = db.get_eleves()
    classes_ecoliers = sorted(set(e['classe'] for e in ecoliers))
    classes_eleves = sorted(set(e['classe'] for e in eleves))
    return render_template('moyenne.html', classes_ecoliers=classes_ecoliers, classes_eleves=classes_eleves, matieres=MATIERES)

@app.route('/calculer_moyenne', methods=['POST'])
def calculer_moyenne():
    d = request.json
    classe, pp_nom, diviseur = d['classe'], d['pp_nom'], int(d['diviseur'])
    students = [s for s in (db.get_ecoliers() if classe in ['maternelle','CI','CP','CE1','CE2','CM1','CM2'] else db.get_eleves()) if s['classe'] == classe]
    matieres_oblig = [m for m in MATIERES if m not in (['EPS','Espagnol'] if classe in ['6ième','5ième'] else [])]
    resultats, manquants = [], []
    for st in students:
        notes = db.get_student_notes(st['id'], 'ecolier' if classe in ['maternelle','CI','CP','CE1','CE2','CM1','CM2'] else 'eleve')
        notes_dict = {n['matiere']: float(n['note']) for n in notes if n['matiere'] in matieres_oblig}
        if len(notes_dict) != len(matieres_oblig):
            manquants.append(f"{st['nom']} {st['prenoms']}"); continue
        moy = round(sum(notes_dict.values()) / diviseur, 1)
        resultats.append({'nom': st['nom'], 'prenoms': st['prenoms'], 'moyenne': moy, 'rang': 0})
    resultats = sorted(resultats, key=lambda x: x['moyenne'], reverse=True)
    for i, r in enumerate(resultats, 1): r['rang'] = i
    return jsonify(success=True, resultats=resultats, effectif=len(resultats), classe=classe, pp_nom=pp_nom, diviseur=diviseur, notes_manquantes=manquants)

@app.route('/export_moyennes', methods=['POST'])
def export_moyennes():
    d = request.json
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Moyennes"
    headers = ['Nom', 'Prénoms', 'Moyenne', 'Rang']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col); c.value = h; c.font = Font(bold=True)
        c.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for r, res in enumerate(d['resultats'], 2):
        ws.cell(row=r, column=1).value = res['nom']
        ws.cell(row=r, column=2).value = res['prenoms']
        ws.cell(row=r, column=3).value = res['moyenne']
        ws.cell(row=r, column=4).value = res['rang']
    io_bytes = io.BytesIO(); wb.save(io_bytes); io_bytes.seek(0)
    return send_file(io_bytes, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'moyennes_{d["classe"]}_{datetime.now():%Y%m%d_%H%M%S}.xlsx')

# ---------- ADMINISTRATIF ----------
@app.route('/administratif')
@require_auth
def administratif():
    return render_template('administratif.html')

@app.route('/api/bilan_admin')
@require_auth
def api_bilan_admin():
    total_du = db.get_total_scolarite_due()
    total_payé = 0
    for el in db.get_eleves():   total_payé += db.get_total_paid(el)
    for ec in db.get_ecoliers(): total_payé += db.get_total_paid(ec)
    return jsonify(total_scolarite_due=total_du, total_scolarite_payee=total_payé)

@app.route('/calculer_admin', methods=['POST'])
def calculer_admin():
    d = request.json
    total_ens = sum(float(x['montant']) for x in d['enseignants'])
    frais_div = float(d['frais_divers'])
    total_payé = float(d['total_scolarite_payee'])
    restant = total_payé - total_ens - frais_div
    resultat = {'total_enseignants': total_ens, 'total_frais': frais_div,
                'total_paye': total_payé, 'restant': restant}
    db.save_admin(resultat, d['enseignants'], frais_div, total_payé)
    return jsonify(success=True, resultat=resultat)

@app.route('/export_bilan_complet', methods=['POST'])
@require_auth
def export_bilan_complet():
    d = request.json
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Bilan"
    lignes = [('Total scolarité DUE',   d['total_due']),
              ('Total scolarité PAYÉE', d['total_payee']),
              ('Total enseignants',     d['total_ens']),
              ('Frais divers',          d['frais_div']),
              ('RESTANT',               d['restant'])]
    for idx, (lib, mt) in enumerate(lignes, 1):
        ws.cell(row=idx, column=1).value = lib
        ws.cell(row=idx, column=2).value = mt
    io_bytes = io.BytesIO(); wb.save(io_bytes); io_bytes.seek(0)
    return send_file(io_bytes, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'bilan_administratif_{datetime.now():%Y%m%d_%H%M%S}.xlsx')

# ---------- SAUVEGARDE ----------
@app.route('/sauvegarde')
@require_auth
def sauvegarde():
    stats = {
        'ecoliers': len(db.get_ecoliers()),
        'eleves': len(db.get_eleves()),
        'notes': len(db.get_notes())
    }
    return render_template('sauvegarde.html', stats=stats)

@app.route('/export_excel')
@require_auth
def export_excel():
    wb = openpyxl.Workbook()
    # Écoliers
    ws1 = wb.active
    ws1.title = "Écoliers"
    ws1.append(["ID", "Nom", "Prénoms", "Classe", "Montant", "Téléphone"])
    for e in db.get_ecoliers():
        ws1.append([e['id'], e['nom'], e['prenoms'], e['classe'], e.get('montant_scolarite', 0), e.get('telephone', '')])
    # Élèves
    ws2 = wb.create_sheet("Élèves")
    ws2.append(["ID", "Nom", "Prénoms", "Classe", "Montant", "Téléphone"])
    for e in db.get_eleves():
        ws2.append([e['id'], e['nom'], e['prenoms'], e['classe'], e.get('montant_scolarite', 0), e.get('telephone', '')])
    # Notes
    ws3 = wb.create_sheet("Notes")
    ws3.append(["Élève", "Classe", "Matière", "Note", "Date"])
    for n in db.get_notes():
        ws3.append([n['student_id'], n['classe'], n['matiere'], n['note'], n['date']])

    io_bytes = io.BytesIO()
    wb.save(io_bytes)
    io_bytes.seek(0)
    return send_file(io_bytes, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'ecole_export_{datetime.now():%Y%m%d_%H%M%S}.xlsx')

@app.route('/import_excel', methods=['GET', 'POST'])
@require_auth
def import_excel():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename.endswith('.xlsx'):
            flash("Fichier invalide (besoin .xlsx)", "error")
            return redirect(url_for('import_excel'))

        wb = openpyxl.load_workbook(file)
        # Écoliers
        if "Écoliers" in wb.sheetnames:
            ws = wb["Écoliers"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[1]: continue
                db.add_ecolier({
                    "nom": str(row[1]).strip(),
                    "prenoms": str(row[2]).strip(),
                    "classe": str(row[3]).strip(),
                    "montant_scolarite": str(row[4] or 0),
                    "telephone": str(row[5] or "")
                })
        # Élèves
        if "Élèves" in wb.sheetnames:
            ws = wb["Élèves"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[1]: continue
                db.add_eleve({
                    "nom": str(row[1]).strip(),
                    "prenoms": str(row[2]).strip(),
                    "classe": str(row[3]).strip(),
                    "montant_scolarite": str(row[4] or 0),
                    "telephone": str(row[5] or "")
                })
        # Notes
        if "Notes" in wb.sheetnames:
            ws = wb["Notes"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                note = float(row[3] or 0)
                if not (0 <= note <= 99): continue
                db.add_note(
                    student_id=int(row[0]),
                    student_type='ecolier' if str(row[1]) in ['maternelle','CI','CP','CE1','CE2','CM1','CM2'] else 'eleve',
                    classe=str(row[1]),
                    matiere=str(row[2]),
                    note=note
                )
        flash("Import réussi !", "success")
        return redirect(url_for('sauvegarde'))
    return render_template('import_excel.html')

# ---------- DEMO ----------
@app.route('/sauvegarde_demo')
def sauvegarde_demo():
    flash('Module en construction', 'info')
    return redirect(url_for('accueil'))

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port, debug=False)
